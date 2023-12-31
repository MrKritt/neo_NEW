import sys
import logging
from datetime import datetime
from domain import use_cases
from domain import repositories
from data_sources.excel_adapter import ExcelAdapter
from data_sources.neosintez_adapter import Neosintez
from utilities import Utilities
from openpyxl import load_workbook


DEBUG = False

MODES = ["notification", "delivery_order", "storage", "appius"]

if __name__ == "__main__":
    if DEBUG:
        mode = "appius"
        config_file_name_suffix = mode
    elif len(sys.argv) != 3 and not DEBUG:
        raise EnvironmentError("Two arguments expected: mode and config suffix")
    else:
        mode = sys.argv[1]
        config_file_name_suffix = sys.argv[2]

    if mode not in MODES:
        raise EnvironmentError(
            f'Invalid mode {mode}. Only {",".join(MODES)} are available'
        )

    config = Utilities.read_config(mode)
    if DEBUG:
        # if debug mode enabled set test files directory
        config["files_directory"] = "test_data/"
    mapping_data = Utilities.mapping_data()
    url = config["url"]
    logs_path = config["logs_path"]
    logging.basicConfig(
        format="%(asctime)s : %(levelname)s : %(message)s",
        level=logging.INFO,
        handlers=[
            logging.FileHandler(
                logs_path
                + datetime.now().strftime("%Y-%m-%d")
                + f"_{config_file_name_suffix}.log"
            ),
            logging.StreamHandler(),
        ],
    )
    logging.info(f"Start {mode} {config_file_name_suffix}")

    ### Дополнительное форматирование столбца Подрядчик + Код
    if mode == "delivery_order":
        base_file_path = config["files_directory"]
        full_file_path = base_file_path + "ЗаказыНаДоставку.xlsx"

        wb = load_workbook(full_file_path)
        ws = wb.active

        for i in range(1, ws.max_column + 1):
            column_name = ws.cell(1, i).value
            if column_name == "Документ заказа.Подрядчик":
                org_name_column = i
            if column_name == "Документ заказа.Подрядчик.Код":
                org_code_column = i

        if org_name_column and org_code_column:
            last_column = ws.max_column + 1
            ws.cell(1, last_column).value = "Документ заказа.Подрядчик, Код"

            for i in range(2, ws.max_row + 1):
                organization = ws.cell(i, org_name_column).value
                code = ws.cell(i, org_code_column).value
                if organization and code:
                    ws.cell(i, last_column).value = (
                        ws.cell(i, org_name_column).value
                        + ", "
                        + ws.cell(i, org_code_column).value
                    )

            wb.save(full_file_path)
            wb.close()
        else:
            logging.error("Could not found specified columns")
            raise ValueError("Check for columns")

    target_adapter = Neosintez(url, config, mapping_data)
    new_data_adapter = ExcelAdapter(
        mode=mode,
        files_directory=config["files_directory"],
        suffix=config["file_suffix"],
        mapping_data=mapping_data,
        key_columns=config["key_columns"],
        key_column_name=config["key_column_name"],
        from_one_file=config.get("from_one_file"),
        filter_column_name_for_one_file=config.get("filter_column_name_for_one_file"),
    )

    one_file = config.get("from_one_file")

    try:
        construction_repository = repositories.ConstructionRepository(target_adapter)
        constructions = construction_repository.get()
        logging.info(f"Total constructions {len(constructions)}")

        for construction in constructions:
            if mode == "notification" and construction.name in [
                "Этилен",
                "Полиэтилен",
                "Восстановление",
                "Опытно-производственная площадка для извлечения лития из подземных вод ЯНГКМ",
                "ПС 110кВ Маччоба",
                "Нефтепровод от КП-8 до УПН",
            ]:
                continue
            elif mode == "delivery_order" and construction.name in [
                "Восстановление",
                "Опытно-производственная площадка для извлечения лития из подземных вод ЯНГКМ",
                "ПС 110кВ Маччоба",
                "Нефтепровод от КП-8 до УПН",
            ]:
                continue
            elif (
                mode == "storage"
                and construction.name != "КС Расширение. Узлы подключения"
            ):
                continue
            else:
                try:
                    logging.info(construction.name)
                    item_repository = repositories.ItemRepository(
                        construction=construction,
                        target_adapter=target_adapter,
                        input_adapter=new_data_adapter,
                        mode=mode,
                        key_column_name=config["key_column_name"],
                        subobject_column_name=config["subobject_column_name"],
                        name_column_name=config["name_column_name"],
                        group_by_column_name=config["group_by_column_name"],
                        save_skipped=config.get("save_skipped") is True,
                        one_root_mode=config.get("one_root_mode") is True,
                    )
                    task = use_cases.IntegrateByModeConstruction(item_repository)
                    task.execute()
                    if not one_file:
                        new_data_adapter.finish()
                except Exception as e:
                    print(e)
                    logging.exception("Exception occurred")

        if one_file:
            new_data_adapter.finish()

    finally:
        target_adapter.close()
        logging.info("Session is closed")
